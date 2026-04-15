"""
export_excel.py
───────────────
Converts oreilly_raw.json into a rich multi-sheet Excel workbook.

Sheets
------
  📋 All Items         – every item, one row each (flat)
  🗂  My Playlists      – grouped by personal playlist
  ⭐ Expert Lists      – grouped by expert/curated playlist
  🎓 Learning Paths    – grouped by learning path
  📊 Summary           – counts, type breakdown, duplicate report

Features
--------
  • Hyperlinks on every URL cell
  • Content-type colour coding (book=blue, video=purple, course=green, …)
  • Bold + frozen header row
  • Auto-sized columns (capped at 80 chars)
  • Duplicate tracker (same URL in multiple collections)
  • Summary pivot by content type

Usage
-----
  python export_excel.py
  python export_excel.py --in ../output/oreilly_raw.json --out ../output/oreilly_export.xlsx
"""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ─── colour palette ───────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)

ROW_FILLS = {
    "book":          PatternFill("solid", fgColor="DDEEFF"),  # light blue
    "video":         PatternFill("solid", fgColor="EDE7F6"),  # light purple
    "course":        PatternFill("solid", fgColor="E8F5E9"),  # light green
    "live-training": PatternFill("solid", fgColor="FFF8E1"),  # light amber
    "sandbox":       PatternFill("solid", fgColor="FCE4EC"),  # light pink
    "article":       PatternFill("solid", fgColor="F3E5F5"),  # light lilac
    "unknown":       PatternFill("solid", fgColor="F5F5F5"),  # light grey
}

SECTION_FILL = PatternFill("solid", fgColor="CFE2F3")   # section header rows

THIN_BORDER_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER      = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE,
)

LINK_FONT = Font(color="1155CC", underline="single")


# ─── helpers ──────────────────────────────────────────────────────────────────

def _apply_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _style_row(ws, row_idx: int, ctype: str, n_cols: int) -> None:
    fill = ROW_FILLS.get(ctype, ROW_FILLS["unknown"])
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=False, vertical="center")


def _auto_width(ws, max_width: int = 80) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _hyperlink(cell, url: str, label: str = "") -> None:
    if not url:
        return
    cell.value    = label or url
    cell.hyperlink = url
    cell.font     = LINK_FONT


def _section_header(ws, text: str, n_cols: int) -> None:
    """Insert a coloured section-separator row."""
    ws.append([text] + [""] * (n_cols - 1))
    row_idx = ws.max_row
    ws.merge_cells(start_row=row_idx, start_column=1,
                   end_row=row_idx,   end_column=n_cols)
    cell = ws.cell(row=row_idx, column=1)
    cell.fill  = SECTION_FILL
    cell.font  = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 20


# ─── flat row builder ─────────────────────────────────────────────────────────

ALL_HEADERS = [
    "collection_name", "collection_type", "collection_url",
    "position", "sub_list_parent",
    "item_title", "content_type", "author", "duration",
    "item_url",
]


def collections_to_rows(data: dict) -> list[dict]:
    """Flatten all collections into one list of row dicts."""
    rows = []
    for section_key, section_label in [
        ("playlists",        "my_playlist"),
        ("expert_playlists", "expert_playlist"),
        ("learning_paths",   "learning_path"),
    ]:
        for coll in data.get(section_key, []):
            for item in coll.get("items", []):
                rows.append({
                    "collection_name": coll["name"],
                    "collection_type": section_label,
                    "collection_url":  coll.get("url", ""),
                    "position":        item.get("position", ""),
                    "sub_list_parent": item.get("sub_list_parent", ""),
                    "item_title":      item.get("title", ""),
                    "content_type":    item.get("content_type", "unknown"),
                    "author":          item.get("author", ""),
                    "duration":        item.get("duration", ""),
                    "item_url":        item.get("url", ""),
                })
    return rows


# ─── sheet builders ───────────────────────────────────────────────────────────

def build_all_items_sheet(ws, rows: list[dict]) -> None:
    _apply_header(ws, ALL_HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_HEADERS))}1"

    for row_dict in rows:
        row_vals = [row_dict.get(h, "") for h in ALL_HEADERS]
        ws.append(row_vals)
        row_idx = ws.max_row
        ctype   = row_dict.get("content_type", "unknown")
        _style_row(ws, row_idx, ctype, len(ALL_HEADERS))

        # Hyperlink the URL column (last column)
        url_col = len(ALL_HEADERS)
        _hyperlink(ws.cell(row=row_idx, column=url_col),
                   row_dict.get("item_url", ""),
                   row_dict.get("item_title", "") or row_dict.get("item_url", ""))

        # Hyperlink the collection_url column (index 2, 0-based → col 3)
        coll_url_col = 3
        _hyperlink(ws.cell(row=row_idx, column=coll_url_col),
                   row_dict.get("collection_url", ""),
                   row_dict.get("collection_name", "") or row_dict.get("collection_url", ""))

    _auto_width(ws)


def build_grouped_sheet(ws, collections: list[dict], source_label: str) -> None:
    HEADERS = ["#", "item_title", "content_type", "author", "duration",
               "sub_list_parent", "item_url"]
    _apply_header(ws, HEADERS)

    for coll in collections:
        _section_header(ws, f"📂  {coll['name']}  ({len(coll['items'])} items)",
                        len(HEADERS))
        for item in coll.get("items", []):
            row_vals = [
                item.get("position", ""),
                item.get("title", ""),
                item.get("content_type", ""),
                item.get("author", ""),
                item.get("duration", ""),
                item.get("sub_list_parent", ""),
                item.get("url", ""),
            ]
            ws.append(row_vals)
            row_idx = ws.max_row
            _style_row(ws, row_idx, item.get("content_type", "unknown"), len(HEADERS))
            _hyperlink(ws.cell(row=row_idx, column=len(HEADERS)),
                       item.get("url", ""),
                       item.get("title", "") or item.get("url", ""))

    _auto_width(ws)


def build_summary_sheet(ws, data: dict, rows: list[dict]) -> None:
    """Summary: totals, type breakdown, cross-collection duplicates."""
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    def bold_row(label, value, row_fill=None):
        ws.append([label, value])
        ri = ws.max_row
        ws.cell(ri, 1).font = Font(bold=True)
        ws.cell(ri, 2).alignment = Alignment(horizontal="right")
        if row_fill:
            for c in range(1, 3):
                ws.cell(ri, c).fill = row_fill

    # ── totals ────────────────────────────────────────────────────────────────
    ws.append(["O'Reilly Export Summary", ""])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")
    ws.append([])

    bold_row("My playlists", len(data.get("playlists", [])), SECTION_FILL)
    bold_row("Expert playlists", len(data.get("expert_playlists", [])))
    bold_row("Learning paths", len(data.get("learning_paths", [])))
    bold_row("Total collections",
             sum(len(data.get(k, [])) for k in ("playlists", "expert_playlists", "learning_paths")),
             HEADER_FILL)
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(ws.max_row, 2).font = Font(bold=True, color="FFFFFF")

    ws.append([])
    bold_row("Total items (incl. sub-list items)", len(rows), SECTION_FILL)

    # ── by content type ───────────────────────────────────────────────────────
    ws.append([])
    ws.append(["Items by content type", "Count"])
    ri = ws.max_row
    for c in range(1, 3):
        cell = ws.cell(ri, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    type_counts = Counter(r.get("content_type", "unknown") for r in rows)
    for ctype, count in type_counts.most_common():
        ws.append([ctype, count])
        _style_row(ws, ws.max_row, ctype, 2)

    # ── by collection type ────────────────────────────────────────────────────
    ws.append([])
    ws.append(["Items by collection type", "Count"])
    ri = ws.max_row
    for c in range(1, 3):
        cell = ws.cell(ri, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    coll_counts = Counter(r.get("collection_type", "") for r in rows)
    for ct, count in coll_counts.most_common():
        ws.append([ct, count])

    # ── duplicates ────────────────────────────────────────────────────────────
    ws.append([])
    ws.append(["Duplicate items (same URL in 2+ collections)", "Collections"])
    ri = ws.max_row
    for c in range(1, 3):
        cell = ws.cell(ri, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    url_to_collections: dict[str, list[str]] = defaultdict(list)
    for r in rows:
        url = r.get("item_url", "")
        coll = r.get("collection_name", "")
        if url and coll:
            if coll not in url_to_collections[url]:
                url_to_collections[url].append(coll)

    dup_count = 0
    for url, colls in url_to_collections.items():
        if len(colls) > 1:
            # Find the title
            title = next((r["item_title"] for r in rows if r.get("item_url") == url), url)
            ws.append([title[:80], ", ".join(colls)])
            ri = ws.max_row
            _hyperlink(ws.cell(ri, 1), url, title[:80])
            dup_count += 1

    if dup_count == 0:
        ws.append(["(no duplicates found)", ""])

    # ── colour legend ─────────────────────────────────────────────────────────
    ws.append([])
    ws.append(["Colour legend", ""])
    ri = ws.max_row
    ws.cell(ri, 1).font = Font(bold=True)

    for ctype, fill in ROW_FILLS.items():
        ws.append([ctype, ""])
        _style_row(ws, ws.max_row, ctype, 2)


# ─── main builder ─────────────────────────────────────────────────────────────

def build_workbook(data: dict, out_path: Path) -> None:
    rows = collections_to_rows(data)

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # All Items
    ws_all = wb.create_sheet("📋 All Items")
    build_all_items_sheet(ws_all, rows)

    # My Playlists
    if data.get("playlists"):
        ws_my = wb.create_sheet("🗂 My Playlists")
        build_grouped_sheet(ws_my, data["playlists"], "my_playlist")

    # Expert Lists
    if data.get("expert_playlists"):
        ws_ex = wb.create_sheet("⭐ Expert Lists")
        build_grouped_sheet(ws_ex, data["expert_playlists"], "expert_playlist")

    # Learning Paths
    if data.get("learning_paths"):
        ws_lp = wb.create_sheet("🎓 Learning Paths")
        build_grouped_sheet(ws_lp, data["learning_paths"], "learning_path")

    # Summary
    ws_sum = wb.create_sheet("📊 Summary")
    build_summary_sheet(ws_sum, data, rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"  ✔  Excel → {out_path.resolve()}")
    print(f"     Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"     Total rows (all items): {len(rows)}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Convert oreilly_raw.json to Excel")
    p.add_argument("--in",  dest="input",  default="../output/oreilly_raw.json")
    p.add_argument("--out", dest="output", default="../output/oreilly_export.xlsx")
    return p.parse_args()


def main():
    args = parse_args()
    in_path  = Path(args.input)
    out_path = Path(args.output)

    if not in_path.exists():
        print(f"❌  Input not found: {in_path.resolve()}")
        return

    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)

    print(f"📥  Loaded raw JSON: {in_path.resolve()}")
    build_workbook(data, out_path)


if __name__ == "__main__":
    main()
