"""
enrich_sheet.py
───────────────
Reads your oreilly_reviews.xlsx, fills in missing data using Playwright,
and saves a new enriched file WITHOUT touching any existing cell values.

For each row it can fill in (only when the cell is currently blank):
  • Link         – GitHub repos + companion websites found on the O'Reilly page
  • Hrs To Complete – scraped from video duration or estimated from page count
  • OReilly_URL  – the O'Reilly page URL (new column, added if missing)

Flow
────
  1. Read xlsx with openpyxl (preserves all existing formatting)
  2. Open Playwright / login via saved auth state (SSO-friendly)
  3. For each row:
       a. If OReilly_URL is already known → quick_read that URL
       b. Else search O'Reilly for the title → get URL → quick_read
  4. Fill only blank cells
  5. Save to output file (never overwrites the source file)

Usage
─────
  python enrich_sheet.py                          # enrich all blank rows
  python enrich_sheet.py --in my_list.xlsx        # custom input
  python enrich_sheet.py --out enriched.xlsx      # custom output
  python enrich_sheet.py --rows 5 10 15           # only these row numbers
  python enrich_sheet.py --skip-search            # only process rows that already have an O'Reilly URL
  python enrich_sheet.py --dry-run                # print what would be filled, don't write
  python enrich_sheet.py --pause 2.0              # slower for flaky connections
  python enrich_sheet.py --resume                 # skip rows already enriched (OReilly_URL filled)
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── local imports ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT / "scraper"))

from auth import ensure_logged_in
from quick_read import quick_read_url, search_oreilly_url

from playwright.sync_api import sync_playwright

# ─── constants ────────────────────────────────────────────────────────────────

DEFAULT_IN  = ROOT / "output" / "oreilly_reviews.xlsx"
DEFAULT_OUT = ROOT / "output" / "oreilly_reviews_enriched.xlsx"

# Columns present in the original file (1-indexed)
COL_SUBJECT  = 1   # A
COL_CONTENT  = 2   # B
COL_AUTHORS  = 3   # C
COL_LINK     = 4   # D
COL_REVISIT  = 5   # E
COL_HRS      = 6   # F
# New column we add
COL_OREILLY  = 7   # G  (OReilly_URL)

# Styling for the new column header
NEW_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
NEW_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

LINK_FONT = Font(color="1155CC", underline="single")


# ─── helpers ──────────────────────────────────────────────────────────────────

def _cell_val(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v not in (None, "", "NaN") else ""


def _is_blank(ws, row: int, col: int) -> bool:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in ("", "nan", "none", "n/a")


def _set_link(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if value.startswith("http"):
        cell.hyperlink = value
        cell.font = LINK_FONT


def _looks_like_oreilly(val: str) -> bool:
    return "learning.oreilly.com" in val or "oreilly.com/library" in val


def _merge_links(existing: str, new_links: list[str]) -> str:
    """Add new links to existing without duplicating."""
    existing_parts = [p.strip() for p in existing.split(",") if p.strip()] if existing else []
    existing_set = {p.lower() for p in existing_parts}
    for lnk in new_links:
        if lnk.lower() not in existing_set:
            existing_parts.append(lnk)
            existing_set.add(lnk.lower())
    return ", ".join(existing_parts)


# ─── xlsx setup ───────────────────────────────────────────────────────────────

def load_workbook_safe(path: Path) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(str(path))
    return wb


def ensure_oreilly_column(ws) -> None:
    """Add the OReilly_URL column header if it doesn't exist."""
    header = ws.cell(row=1, column=COL_OREILLY).value
    if not header or str(header).strip() == "":
        cell = ws.cell(row=1, column=COL_OREILLY)
        cell.value = "OReilly_URL"
        cell.fill  = NEW_HEADER_FILL
        cell.font  = NEW_HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(COL_OREILLY)].width = 55


# ─── main enrichment ──────────────────────────────────────────────────────────

def enrich(args) -> None:
    in_path  = Path(args.input)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not in_path.exists():
        print(f"❌  Input file not found: {in_path.resolve()}")
        sys.exit(1)

    wb = load_workbook_safe(in_path)
    ws = wb.active

    ensure_oreilly_column(ws)

    # Determine which rows to process
    max_row = ws.max_row
    if args.rows:
        # +1 because row 1 is header
        row_indices = [r + 1 for r in args.rows if 1 <= r <= max_row - 1]
    else:
        row_indices = list(range(2, max_row + 1))

    # Filter to rows that actually need work
    needs_work = []
    for ri in row_indices:
        title = _cell_val(ws, ri, COL_CONTENT)
        if not title:
            continue

        missing_link = _is_blank(ws, ri, COL_LINK)
        missing_hrs  = _is_blank(ws, ri, COL_HRS)
        has_oreilly  = not _is_blank(ws, ri, COL_OREILLY)

        if args.resume and has_oreilly and not missing_link and not missing_hrs:
            continue
        if args.skip_search and not has_oreilly:
            continue
        if not missing_link and not missing_hrs and has_oreilly:
            continue  # nothing to do

        needs_work.append(ri)

    print(f"\n📋  Rows to process: {len(needs_work)} / {max_row - 1}")
    if not needs_work:
        print("✅  Nothing to enrich — all rows already complete.")
        return

    if args.dry_run:
        print("🔍  Dry run — rows that would be processed:")
        for ri in needs_work:
            print(f"    row {ri - 1:3d}: {_cell_val(ws, ri, COL_CONTENT)[:60]}")
        return

    # ── launch browser ────────────────────────────────────────────────────────
    results_log: list[dict] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        ctx, page = ensure_logged_in(browser)

        for idx, ri in enumerate(needs_work, 1):
            title   = _cell_val(ws, ri, COL_CONTENT)
            subject = _cell_val(ws, ri, COL_SUBJECT)
            cur_link    = _cell_val(ws, ri, COL_LINK)
            cur_oreilly = _cell_val(ws, ri, COL_OREILLY)
            cur_hrs     = _cell_val(ws, ri, COL_HRS)

            print(f"\n[{idx}/{len(needs_work)}]  Row {ri - 1}: {title[:60]}")

            oreilly_url = cur_oreilly if cur_oreilly else None

            # ── step 1: find O'Reilly URL if missing ──────────────────────────
            if not oreilly_url and not args.skip_search:
                print(f"  🔍  Searching O'Reilly for: {title[:50]}")
                try:
                    oreilly_url = search_oreilly_url(page, title)
                    if oreilly_url:
                        print(f"      Found: {oreilly_url}")
                    else:
                        print(f"      ⚠️  No result found — skipping enrichment")
                except Exception as e:
                    print(f"      ⚠️  Search error: {e}")

            # ── step 2: quick_read the O'Reilly page ──────────────────────────
            enriched = {}
            if oreilly_url and _looks_like_oreilly(oreilly_url):
                try:
                    enriched = quick_read_url(page, oreilly_url, pause=args.pause)
                    print(f"      ⏱️  Hours: {enriched.get('hours')}  "
                          f"🔗  Links: {len(enriched.get('links', []))}")
                except Exception as e:
                    print(f"      ⚠️  Quick read error: {e}")

            # ── step 3: write back (only blanks) ──────────────────────────────
            changed: dict[str, str] = {}

            # OReilly_URL
            if oreilly_url and _is_blank(ws, ri, COL_OREILLY):
                _set_link(ws, ri, COL_OREILLY, oreilly_url)
                changed["OReilly_URL"] = oreilly_url

            # Hours to complete
            if enriched.get("hours") and _is_blank(ws, ri, COL_HRS):
                ws.cell(row=ri, column=COL_HRS).value = enriched["hours"]
                changed["Hrs"] = str(enriched["hours"])

            # Links (GitHub + companion sites)
            new_links = enriched.get("links", [])[:3]  # top 3
            if new_links:
                if _is_blank(ws, ri, COL_LINK):
                    merged = ", ".join(new_links)
                    _set_link(ws, ri, COL_LINK, merged)
                    changed["Link"] = merged
                else:
                    # Append new links not already present
                    merged = _merge_links(cur_link, new_links)
                    if merged != cur_link:
                        ws.cell(row=ri, column=COL_LINK).value = merged
                        changed["Link (+appended)"] = merged

            if changed:
                print(f"      ✔  Updated: {', '.join(changed.keys())}")
            else:
                print(f"      —  No new data found")

            results_log.append({
                "row": ri - 1,
                "title": title,
                "oreilly_url": oreilly_url,
                **enriched,
                "changed": changed,
            })

            # Save incrementally every 5 rows
            if idx % 5 == 0:
                wb.save(str(out_path))
                print(f"  💾  Saved progress → {out_path.name}")

            time.sleep(0.3)  # brief pause between rows

        ctx.close()
        browser.close()

    # Final save
    wb.save(str(out_path))

    # Save log
    log_path = out_path.parent / "enrich_log.json"
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(results_log, f, ensure_ascii=False, indent=2, default=str)

    # ── summary ───────────────────────────────────────────────────────────────
    filled_links = sum(1 for r in results_log if "Link" in r.get("changed", {}))
    filled_hrs   = sum(1 for r in results_log if "Hrs"  in r.get("changed", {}))
    filled_urls  = sum(1 for r in results_log if "OReilly_URL" in r.get("changed", {}))

    print(f"\n{'═'*55}")
    print(f"✅  Enrichment complete")
    print(f"    OReilly URLs added:  {filled_urls}")
    print(f"    Links filled:        {filled_links}")
    print(f"    Hours filled:        {filled_hrs}")
    print(f"    Output:  {out_path.resolve()}")
    print(f"    Log:     {log_path.resolve()}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Enrich oreilly_reviews.xlsx with scraped metadata",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--in",  dest="input",  default=str(DEFAULT_IN),
                   help=f"Input xlsx (default: {DEFAULT_IN})")
    p.add_argument("--out", dest="output", default=str(DEFAULT_OUT),
                   help=f"Output xlsx (default: {DEFAULT_OUT})")
    p.add_argument("--rows",        nargs="+", type=int,
                   help="Only process these row numbers (1-indexed, ignoring header)")
    p.add_argument("--skip-search", action="store_true",
                   help="Skip rows without an existing OReilly_URL (no searching)")
    p.add_argument("--dry-run",     action="store_true",
                   help="Show what would be processed without making changes")
    p.add_argument("--resume",      action="store_true",
                   help="Skip rows that already have OReilly_URL + Link + Hrs filled")
    p.add_argument("--pause",       type=float, default=1.5,
                   help="Seconds to wait after loading each page (default 1.5)")
    return p.parse_args()


if __name__ == "__main__":
    enrich(parse_args())
