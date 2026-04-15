"""
build_review_list.py
────────────────────
Reads oreilly_raw.json (output of the playlist scraper), visits each
item's landing page, and builds an Excel sheet in the same format as
oreilly_reviews.xlsx — automatically filling in the columns you were
completing by hand.

Output columns (matching oreilly_reviews.xlsx + new ones)
─────────────────────────────────────────────────────────
  A  Subject           ← playlist / collection name (= category)
  B  Content           ← book / video title
  C  Authors           ← scraped from landing page
  D  Link              ← GitHub repos + companion sites from description
  E  Revisit           ← left blank (you fill this in)
  F  Hrs To Complete   ← scraped (video: exact; book: estimated from pages)
  G  OReilly_URL       ← the O'Reilly page URL
  H  TOC               ← chapter / section titles, one per line
  I  Tags              ← O'Reilly topic tags

Run modes
─────────
  python build_review_list.py                       # full run
  python build_review_list.py --resume              # skip items already in output
  python build_review_list.py --limit 10            # only first N items
  python build_review_list.py --collection "ML"     # only one collection
  python build_review_list.py --pause 2.0           # slower for flaky connections

Also works on your existing oreilly_reviews.xlsx to enrich missing rows:
  python build_review_list.py --mode enrich \\
         --in output/oreilly_reviews.xlsx \\
         --out output/oreilly_reviews_enriched.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT / "scraper"))

from auth import ensure_logged_in
from toc_scraper import scrape_landing_page, search_oreilly
from playwright.sync_api import sync_playwright

# ─── paths ────────────────────────────────────────────────────────────────────

RAW_JSON    = ROOT / "output" / "oreilly_raw.json"
OUT_XLSX    = ROOT / "output" / "oreilly_reviews_built.xlsx"
PROGRESS_F  = ROOT / "output" / "build_progress.json"

# ─── column map ───────────────────────────────────────────────────────────────

C = {
    "subject":  1,   # A
    "content":  2,   # B
    "authors":  3,   # C
    "link":     4,   # D
    "revisit":  5,   # E
    "hrs":      6,   # F
    "url":      7,   # G
    "toc":      8,   # H
    "tags":     9,   # I
}

HEADERS = [
    "Subject", "Content", "Authors", "Link",
    "Revisit", "Hrs To Complete", "OReilly_URL", "TOC", "Tags",
]

COL_WIDTHS = {
    "A": 22,   # Subject
    "B": 55,   # Content
    "C": 35,   # Authors
    "D": 55,   # Link
    "E": 10,   # Revisit
    "F": 16,   # Hrs To Complete
    "G": 58,   # OReilly_URL
    "H": 60,   # TOC
    "I": 35,   # Tags
}

# Colours matching oreilly_reviews.xlsx plain style (no header fill — matches original)
HEADER_FONT  = Font(name="Calibri", size=11, bold=True)
DATA_FONT    = Font(name="Calibri", size=11)
LINK_FONT    = Font(name="Calibri", size=11, color="1155CC", underline="single")

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(bottom=THIN)

# Category → fill colour (light pastels, one per subject group)
SUBJECT_FILLS = {
    "api and microservices":        "FFF2CC",
    "aws":                          "E2EFDA",
    "c++":                          "DDEBF7",
    "data science":                 "FCE4D6",
    "deep learning":                "EAD1DC",
    "interviews":                   "D9D9D9",
    "java":                         "FFE699",
    "machine learning":             "C6EFCE",
    "machine learning engineering": "C6EFCE",
    "python":                       "BDD7EE",
    "software engineering":         "F4CCCC",
    "spark":                        "FFD966",
    "trading":                      "B7E1CD",
    "webdevelopment":               "D0E0E3",
}


# ─── workbook helpers ─────────────────────────────────────────────────────────

def new_workbook() -> tuple[openpyxl.Workbook, object]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "oreilly_reviews"

    # Header row
    ws.append(HEADERS)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c)
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = Border(bottom=Side(style="medium", color="888888"))

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    return wb, ws


def _subject_fill(subject: str) -> PatternFill | None:
    key = subject.strip().lower()
    rgb = SUBJECT_FILLS.get(key)
    if rgb:
        return PatternFill("solid", fgColor=rgb)
    return None


def write_row(ws, row_idx: int, data: dict) -> None:
    """Write one data row with formatting."""
    subject  = data.get("subject", "")
    content  = data.get("content", "")
    authors  = data.get("authors", "")
    link     = data.get("link", "")
    revisit  = data.get("revisit", "")
    hrs      = data.get("hrs")
    oreilly  = data.get("oreilly_url", "")
    toc_list = data.get("toc", [])
    tags     = data.get("tags", [])

    toc_str  = "\n".join(toc_list) if toc_list else ""
    tags_str = ", ".join(tags)     if tags     else ""

    values = [subject, content, authors, link, revisit, hrs, oreilly, toc_str, tags_str]
    fill = _subject_fill(subject)

    for c, val in enumerate(values, 1):
        cell = ws.cell(row_idx, c)
        cell.value     = val
        cell.font      = DATA_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (C["toc"],)))
        cell.border    = Border(bottom=THIN)
        if fill:
            cell.fill = fill

    # Hyperlink OReilly_URL
    if oreilly:
        cell = ws.cell(row_idx, C["url"])
        cell.hyperlink = oreilly
        cell.font      = LINK_FONT

    # Hyperlink first GitHub link in Link cell
    if link and link.startswith("http"):
        first_url = link.split(",")[0].strip()
        if first_url.startswith("http"):
            cell = ws.cell(row_idx, C["link"])
            cell.hyperlink = first_url
            cell.font      = LINK_FONT

    # Reasonable row height (taller if TOC present)
    ws.row_dimensions[row_idx].height = min(15 + len(toc_list) * 14, 200)


def save_wb(wb: openpyxl.Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ─── progress (resume support) ────────────────────────────────────────────────

def load_progress() -> dict:
    if PROGRESS_F.exists():
        with open(PROGRESS_F, encoding="utf-8") as f:
            return json.load(f)
    return {}   # keyed by oreilly_url → result dict


def save_progress(prog: dict) -> None:
    with open(PROGRESS_F, "w", encoding="utf-8") as f:
        json.dump(prog, f, ensure_ascii=False, indent=2, default=str)


# ─── build from raw JSON ──────────────────────────────────────────────────────

def items_from_raw(raw: dict, collection_filter: str | None) -> list[dict]:
    """
    Flatten oreilly_raw.json into a list of dicts with keys:
      subject, content, oreilly_url (if known), collection_type
    """
    items = []
    for section_key, section_label in [
        ("playlists",        "my_playlist"),
        ("expert_playlists", "expert_playlist"),
        ("learning_paths",   "learning_path"),
    ]:
        for coll in raw.get(section_key, []):
            subject = coll.get("name", "")
            if collection_filter and collection_filter.lower() not in subject.lower():
                continue
            for item in coll.get("items", []):
                items.append({
                    "subject":         subject,
                    "content":         item.get("title", ""),
                    "oreilly_url":     item.get("url", ""),
                    "collection_type": section_label,
                })
    return items


# ─── enrich mode: fill blanks in existing xlsx ───────────────────────────────

def _is_blank(v) -> bool:
    return v is None or str(v).strip().lower() in ("", "nan", "none", "n/a")


def enrich_existing(args, page) -> None:
    in_path  = Path(args.input)
    out_path = Path(args.output)

    wb = openpyxl.load_workbook(str(in_path))
    ws = wb.active

    # Ensure new columns exist
    existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for col_name, col_idx in [("OReilly_URL", C["url"]), ("TOC", C["toc"]), ("Tags", C["tags"])]:
        if col_name not in existing_headers:
            cell = ws.cell(1, col_idx)
            cell.value = col_name
            cell.font  = HEADER_FONT
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(
                get_column_letter(col_idx), 40)

    prog = load_progress() if args.resume else {}

    total = ws.max_row - 1
    changed = 0

    for ri in range(2, ws.max_row + 1):
        title   = str(ws.cell(ri, C["content"]).value or "").strip()
        subject = str(ws.cell(ri, C["subject"]).value or "").strip()
        if not title:
            continue

        cur_url  = str(ws.cell(ri, C["url"]).value or "").strip()
        cur_link = str(ws.cell(ri, C["link"]).value or "").strip()
        cur_hrs  = ws.cell(ri, C["hrs"]).value
        cur_toc  = str(ws.cell(ri, C["toc"]).value or "").strip()

        need_url  = _is_blank(cur_url)
        need_link = _is_blank(cur_link) or cur_link.lower() in ("video watched",)
        need_hrs  = _is_blank(cur_hrs)
        need_toc  = _is_blank(cur_toc)

        if not any([need_url, need_link, need_hrs, need_toc]):
            print(f"  row {ri-1:3d}: ✓ complete — {title[:50]}")
            continue

        print(f"\n  row {ri-1:3d}: [{subject}] {title[:55]}")

        # Find URL
        oreilly_url = cur_url if not _is_blank(cur_url) else None
        if not oreilly_url:
            if not _is_blank(cur_link) and "oreilly.com" in cur_link:
                oreilly_url = cur_link.split(",")[0].strip()
            elif not args.skip_search:
                print(f"    🔍  Searching …")
                oreilly_url = search_oreilly(page, title)
                if oreilly_url:
                    print(f"    ✔  {oreilly_url}")

        if not oreilly_url:
            print(f"    ⚠️  No URL found — skipping")
            continue

        # Check progress cache
        scraped = prog.get(oreilly_url)
        if not scraped:
            scraped = scrape_landing_page(page, oreilly_url, pause=args.pause)
            prog[oreilly_url] = scraped
            save_progress(prog)

        # Write back — only blanks
        if need_url:
            cell = ws.cell(ri, C["url"])
            cell.value     = oreilly_url
            cell.hyperlink = oreilly_url
            cell.font      = LINK_FONT

        if need_hrs and scraped.get("hours"):
            ws.cell(ri, C["hrs"]).value = scraped["hours"]

        if need_link and scraped.get("links"):
            new_link = ", ".join(scraped["links"][:3])
            cell = ws.cell(ri, C["link"])
            cell.value     = new_link
            cell.hyperlink = scraped["links"][0]
            cell.font      = LINK_FONT

        if need_toc and scraped.get("toc"):
            cell = ws.cell(ri, C["toc"])
            cell.value     = "\n".join(scraped["toc"])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[ri].height = min(15 + len(scraped["toc"]) * 14, 200)

        if not ws.cell(ri, C["tags"]).value and scraped.get("tags"):
            ws.cell(ri, C["tags"]).value = ", ".join(scraped["tags"])

        changed += 1
        print(f"    ✔  hrs={scraped.get('hours')}  "
              f"links={len(scraped.get('links',[]))}  "
              f"toc={len(scraped.get('toc',[]))} chapters")

        if changed % 5 == 0:
            save_wb(wb, out_path)
            print(f"    💾  Progress saved ({changed} rows enriched)")

    save_wb(wb, out_path)
    print(f"\n✅  Enrich done — {changed}/{total} rows updated → {out_path.resolve()}")


# ─── build mode: build fresh from raw JSON ────────────────────────────────────

def build_from_raw(args, page) -> None:
    raw_path = Path(args.raw_json)
    out_path = Path(args.output)

    if not raw_path.exists():
        print(f"❌  oreilly_raw.json not found: {raw_path.resolve()}")
        print("    Run `python run_pipeline.py --steps scrape` first.")
        sys.exit(1)

    with open(raw_path, encoding="utf-8") as f:
        raw = json.load(f)

    items = items_from_raw(raw, args.collection)
    if args.limit:
        items = items[:args.limit]

    total = len(items)
    print(f"\n📋  {total} items to process")

    prog = load_progress() if args.resume else {}

    wb, ws = new_workbook()
    row_idx = 2

    for i, item in enumerate(items, 1):
        title       = item["content"]
        subject     = item["subject"]
        oreilly_url = item.get("oreilly_url", "")

        print(f"\n[{i:3d}/{total}]  [{subject}]  {title[:55]}")

        # ── find URL if not already known ────────────────────────────────────
        if not oreilly_url or "oreilly.com" not in oreilly_url:
            if not args.skip_search:
                print(f"         🔍  Searching …")
                oreilly_url = search_oreilly(page, title) or ""
                if oreilly_url:
                    print(f"         ✔  {oreilly_url}")
                else:
                    print(f"         ⚠️  Not found on O'Reilly — writing stub row")
            else:
                print(f"         ⚠️  No URL, skipping (--skip-search mode)")
                oreilly_url = ""

        # ── scrape landing page ───────────────────────────────────────────────
        scraped: dict = {}
        if oreilly_url:
            scraped = prog.get(oreilly_url, {})
            if not scraped:
                scraped = scrape_landing_page(page, oreilly_url, pause=args.pause)
                prog[oreilly_url] = scraped
                save_progress(prog)

            print(f"         ⏱  hrs={scraped.get('hours')}  "
                  f"links={len(scraped.get('links',[]))}  "
                  f"toc={len(scraped.get('toc',[]))} items")

        # ── merge: prefer scraped data, fall back to what was in raw JSON ─────
        row_data = {
            "subject":     subject,
            "content":     scraped.get("title") or title,
            "authors":     scraped.get("authors", ""),
            "link":        ", ".join(scraped.get("links", [])[:3]),
            "revisit":     "",
            "hrs":         scraped.get("hours"),
            "oreilly_url": oreilly_url,
            "toc":         scraped.get("toc", []),
            "tags":        scraped.get("tags", []),
        }

        write_row(ws, row_idx, row_data)
        row_idx += 1

        # Incremental save every 10 rows
        if i % 10 == 0:
            save_wb(wb, out_path)
            print(f"         💾  Saved ({i} rows written)")

    save_wb(wb, out_path)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    save_wb(wb, out_path)
    print(f"\n✅  Built {row_idx - 2} rows → {out_path.resolve()}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Build oreilly_reviews sheet from scraped playlists",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--mode", choices=["build", "enrich"], default="build",
                   help="'build' = new sheet from oreilly_raw.json | "
                        "'enrich' = fill blanks in existing xlsx")

    # build-mode options
    p.add_argument("--raw-json",   default=str(RAW_JSON))
    p.add_argument("--collection", default=None,
                   help="Only process items from this collection name")
    p.add_argument("--limit", type=int, default=0,
                   help="Stop after N items (0 = all)")

    # enrich-mode options
    p.add_argument("--in",  dest="input",
                   default=str(ROOT / "output" / "oreilly_reviews.xlsx"))

    # shared
    p.add_argument("--out", dest="output", default=str(OUT_XLSX))
    p.add_argument("--pause",       type=float, default=1.5)
    p.add_argument("--resume",      action="store_true",
                   help="Skip items already scraped (uses build_progress.json)")
    p.add_argument("--skip-search", action="store_true",
                   help="Don't search O'Reilly for items without a URL")
    return p.parse_args()


def main():
    args = parse_args()

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        ctx, page = ensure_logged_in(browser)

        if args.mode == "enrich":
            enrich_existing(args, page)
        else:
            build_from_raw(args, page)

        ctx.close()
        browser.close()


if __name__ == "__main__":
    main()
